# -*- coding: utf-8 -*-
"""
코스피/코스닥/선물 투자자별 매매동향(수급) 일일 수집기

  - 출처: 네이버 금융 투자자별 매매동향 (finance.naver.com/sise/investorDealTrendDay.naver)
          sosok=01 코스피 현물, 02 코스닥 현물, 03 선물(KOSPI200)
  - 지수: 네이버 모바일 API (m.stock.naver.com/api/index/{code}/price)
  - 단위: 순매수 금액, 억원
  - 저장: data/*.csv (마스터) -> 수급동향.xlsx + dashboard_data.js 재생성
  - 실행: python collect.py            # 증분 수집(최근 10거래일 갱신)
          python collect.py --backfill # 2026-01-02부터 전체 수집

주의: 수급동향.xlsx 와 dashboard_data.js 는 매 실행마다 CSV 마스터로부터
      전체 재생성되므로 직접 수정하면 안 된다.
"""
import io
import os
import sys
import time
import json
import logging
from datetime import datetime, date, timedelta

import requests
import pandas as pd

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE, "data")
LOG_DIR = os.path.join(BASE, "logs")
XLSX_PATH = os.path.join(BASE, "수급동향.xlsx")
DASH_JS_PATH = os.path.join(BASE, "dashboard_data.js")

BACKFILL_START = date(2026, 1, 2)
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
MAX_PAGES = 60          # 안전장치 (10거래일/페이지)
REQ_DELAY = 0.3

# (마켓키, sosok, 한글명, 지수코드)
MARKETS = [
    ("kospi",   "01", "코스피", "KOSPI"),
    ("kosdaq",  "02", "코스닥", "KOSDAQ"),
    ("futures", "03", "선물",   "KPI200"),
]

FLOW_COLS = [
    "individual", "foreign", "inst_total", "fin_invest", "insurance",
    "invest_trust", "bank", "other_fin", "pension", "other_corp",
]
FLOW_COLS_KR = [
    "개인", "외국인", "기관계", "금융투자", "보험",
    "투신(사모)", "은행", "기타금융", "연기금등", "기타법인",
]


def setup_logging():
    os.makedirs(LOG_DIR, exist_ok=True)
    path = os.path.join(LOG_DIR, f"collect_{datetime.now():%Y%m}.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


def fetch(url, retries=3):
    last = None
    for i in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            return r
        except Exception as e:  # noqa: BLE001
            last = e
            time.sleep(1.5 * (i + 1))
    raise RuntimeError(f"요청 실패: {url} ({last})")


def parse_trend_page(html):
    """investorDealTrendDay 페이지 -> [(date, {flow_col: 억원}), ...]"""
    tables = pd.read_html(io.StringIO(html))
    if not tables:
        return []
    t = tables[0]
    t.columns = ["날짜"] + FLOW_COLS
    rows = []
    for _, row in t.iterrows():
        s = str(row["날짜"])
        if len(s) != 8 or s[2] != "." or s[5] != ".":
            continue
        try:
            d = datetime.strptime(s, "%y.%m.%d").date()
        except ValueError:
            continue
        vals = {}
        ok = True
        for c in FLOW_COLS:
            v = row[c]
            if pd.isna(v):
                ok = False
                break
            vals[c] = int(v)
        if ok:
            rows.append((d, vals))
    return rows


# ---------------------------------------------------------------- 한투 수급
# finance.naver 종료(2026-09-10) 대비: 코스피·코스닥 세부 수급은 한국투자증권
# 공식 API(시장별 투자자매매동향 일별, FHPTJ04040000)로 전환(2026-08-28).
# 네이버 값과 4거래일 전수 대조로 일치 확인(단위: 백만원 → /100 = 억원).
# 선물(sosok=03)은 이 TR 미지원이라 당분간 구형 페이지를 유지한다 — 종료 후
# 대체 소스(한투 선물옵션 TR 또는 KRX) 확정 필요.
_HANTU_ISCD = {"01": ("0001", "KSP"), "02": ("1001", "KSQ")}
# FLOW_COLS → 한투 응답 필드(*_ntby_tr_pbmn, 백만원). 튜플이면 합산 —
# 네이버 분류 기준: 투신 = 투신+사모, 기타금융 = 기타금융+종금
# (2026-08-18 코스피 값으로 잔차 0.2억 이내 정합 확인).
_HANTU_COLMAP = {
    "individual": ("prsn_ntby_tr_pbmn",), "foreign": ("frgn_ntby_tr_pbmn",),
    "inst_total": ("orgn_ntby_tr_pbmn",), "fin_invest": ("scrt_ntby_tr_pbmn",),
    "insurance": ("insu_ntby_tr_pbmn",),
    "invest_trust": ("ivtr_ntby_tr_pbmn", "pe_fund_ntby_tr_pbmn"),
    "bank": ("bank_ntby_tr_pbmn",),
    "other_fin": ("etc_orgt_ntby_tr_pbmn", "mrbn_ntby_tr_pbmn"),
    "pension": ("fund_ntby_tr_pbmn",), "other_corp": ("etc_corp_ntby_tr_pbmn",),
}
_hantu_token_cache = {}


def _hantu_cfg():
    import yaml
    root = os.path.dirname(BASE)
    cfg = yaml.safe_load(open(os.path.join(root, "config.yaml"), encoding="utf-8"))
    return cfg["hantu"]


def _hantu_token(cfg):
    """접근토큰 — 파일 캐시(23시간). 한투는 1분당 발급 횟수 제한이 있다."""
    import requests as rq
    cache = os.path.join(BASE, "data", ".hantu_token.json")
    try:
        c = json.load(open(cache, encoding="utf-8"))
        if time.time() - c["ts"] < 23 * 3600:
            return c["token"]
    except Exception:
        pass
    res = rq.post("https://openapi.koreainvestment.com:9443/oauth2/tokenP",
                  json={"grant_type": "client_credentials",
                        "appkey": cfg["api_key"], "appsecret": cfg["secret_key"]},
                  timeout=15).json()
    tok = res["access_token"]
    os.makedirs(os.path.dirname(cache), exist_ok=True)
    json.dump({"token": tok, "ts": time.time()}, open(cache, "w", encoding="utf-8"))
    return tok


def fetch_flows_hantu(sosok, start_date, label):
    """한투 TR 로 start_date~오늘 구간 수급 수집. 실패 시 예외(호출부 폴백)."""
    import requests as rq
    cfg = _hantu_cfg()
    tok = _hantu_token(cfg)
    iscd, tag = _HANTU_ISCD[sosok]
    h = {"authorization": f"Bearer {tok}", "appkey": cfg["api_key"],
         "appsecret": cfg["secret_key"], "tr_id": "FHPTJ04040000", "custtype": "P"}
    # DATE_1 이 '기준일'이고 응답은 그날(포함)부터 과거로 최대 300행 —
    # 기간 파라미터처럼 보이지만 시작일을 넣으면 옛 데이터만 온다(실측).
    today_str = datetime.now().strftime("%Y%m%d")
    params = {"FID_COND_MRKT_DIV_CODE": "U", "FID_INPUT_ISCD": iscd,
              "FID_INPUT_DATE_1": today_str, "FID_INPUT_DATE_2": today_str,
              "FID_INPUT_ISCD_1": tag, "FID_INPUT_ISCD_2": iscd,
              "FID_COND_SCR_DIV_CODE": "16449"}
    r = rq.get("https://openapi.koreainvestment.com:9443"
               "/uapi/domestic-stock/v1/quotations/inquire-investor-daily-by-market",
               headers=h, params=params, timeout=20)
    d = r.json()
    if d.get("rt_cd") != "0":
        raise RuntimeError(f"한투 수급 조회 실패: {d.get('msg1')}")
    out = {}
    for row in d.get("output") or []:
        try:
            dt_ = datetime.strptime(row["stck_bsop_date"], "%Y%m%d").date()
        except (KeyError, ValueError):
            continue
        if dt_ < start_date:
            continue
        vals, ok = {}, True
        for col, fields in _HANTU_COLMAP.items():
            try:
                vals[col] = int(round(sum(float(row[f]) for f in fields) / 100.0))  # 백만원 → 억원
            except (KeyError, TypeError, ValueError):
                ok = False
                break
        if ok:
            out[dt_] = vals
    logging.info("%s 수급 %d일 수집 — 한투 API (%s ~)", label, len(out), start_date)
    return out


def fetch_flows(sosok, start_date, label):
    """start_date까지 수급 수집 — 코스피/코스닥은 한투 API, 실패·선물은 구형 페이지."""
    if sosok in _HANTU_ISCD:
        try:
            return fetch_flows_hantu(sosok, start_date, label)
        except Exception as e:
            logging.warning("%s 한투 수급 실패(%s) — 구형 페이지 폴백", label, e)
    bizdate = datetime.now().strftime("%Y%m%d")
    out = {}
    for page in range(1, MAX_PAGES + 1):
        url = (
            "https://finance.naver.com/sise/investorDealTrendDay.naver"
            f"?bizdate={bizdate}&sosok={sosok}&page={page}"
        )
        r = fetch(url)
        if "charset" not in (r.headers.get("Content-Type") or "").lower():
            r.encoding = "euc-kr"   # 서버가 charset 미제공 시만 (네이버가 UTF-8 전환 중, 2026-08-26)
        rows = parse_trend_page(r.text)
        if not rows:
            break
        for d, vals in rows:
            if d >= start_date:
                out[d] = vals
        if min(d for d, _ in rows) < start_date:
            break
        time.sleep(REQ_DELAY)
    logging.info("%s 수급 %d일 수집 (%s ~)", label, len(out), start_date)
    return out


def fetch_index(code, start_date):
    """지수 일별 종가/등락률 수집 -> {date: (close, change_pct)}"""
    out = {}
    for page in range(1, MAX_PAGES + 1):
        url = f"https://m.stock.naver.com/api/index/{code}/price?pageSize=60&page={page}"
        items = fetch(url).json()
        if not items:
            break
        oldest = None
        for it in items:
            d = datetime.strptime(it["localTradedAt"], "%Y-%m-%d").date()
            oldest = d if oldest is None or d < oldest else oldest
            if d >= start_date:
                close = float(it["closePrice"].replace(",", ""))
                pct = float(it["fluctuationsRatio"].replace(",", ""))
                out[d] = (close, pct)
        if oldest is not None and oldest < start_date:
            break
        time.sleep(REQ_DELAY)
    logging.info("%s 지수 %d일 수집", code, len(out))
    return out


def load_csv(path):
    if os.path.exists(path):
        df = pd.read_csv(path, encoding="utf-8-sig", parse_dates=["date"])
        df["date"] = df["date"].dt.date
        return df
    return pd.DataFrame()


def upsert(old, new_rows, columns):
    """new_rows: {date: dict|tuple} — 같은 날짜는 새 값으로 교체."""
    recs = []
    for d, v in new_rows.items():
        rec = {"date": d}
        if isinstance(v, dict):
            rec.update(v)
        else:
            rec.update(dict(zip(columns, v)))
        recs.append(rec)
    new_df = pd.DataFrame(recs)
    if not old.empty:
        old = old[~old["date"].isin(new_df["date"])]
        new_df = pd.concat([old, new_df], ignore_index=True)
    new_df = new_df.sort_values("date").reset_index(drop=True)
    return new_df


def build_excel(flow_dfs, index_dfs):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    f_base = Font(name="맑은 고딕", size=10)
    f_head = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    fill_head = PatternFill("solid", fgColor="2F5597")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")

    fmt_flow = "[Red]#,##0;[Blue]-#,##0;0"
    fmt_cum = "#,##0;-#,##0;0"
    fmt_idx = "#,##0.00"
    fmt_pct = "+0.00;-0.00;0.00"

    for mkey, _, mname, icode in MARKETS:
        df = flow_dfs[mkey].copy()
        idx = index_dfs[mkey]
        if not idx.empty:
            df = df.merge(idx, on="date", how="left")
        else:
            df["close"] = None
            df["change_pct"] = None
        for c in ["individual", "foreign", "inst_total"]:
            df[f"cum_{c}"] = df[c].cumsum()

        idx_name = {"KOSPI": "코스피", "KOSDAQ": "코스닥", "KPI200": "K200"}[icode]
        headers = (
            ["날짜"] + FLOW_COLS_KR
            + ["개인 누적", "외국인 누적", "기관 누적",
               f"{idx_name} 종가", "등락률(%)"]
        )
        cols = (
            ["date"] + FLOW_COLS
            + ["cum_individual", "cum_foreign", "cum_inst_total", "close", "change_pct"]
        )

        ws = wb.create_sheet(mname)
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = f_head
            c.fill = fill_head
            c.alignment = center
            c.border = border
        for i, (_, row) in enumerate(df.iterrows(), start=2):
            for j, col in enumerate(cols, start=1):
                v = row[col]
                if col == "date":
                    c = ws.cell(row=i, column=j, value=row["date"].strftime("%Y-%m-%d"))
                    c.alignment = center
                elif pd.isna(v):
                    c = ws.cell(row=i, column=j, value=None)
                else:
                    c = ws.cell(row=i, column=j, value=float(v) if col in ("close", "change_pct") else int(v))
                    if col in FLOW_COLS:
                        c.number_format = fmt_flow
                    elif col.startswith("cum_"):
                        c.number_format = fmt_cum
                    elif col == "close":
                        c.number_format = fmt_idx
                    else:
                        c.number_format = fmt_pct
                c.font = f_base
                c.border = border
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df) + 1}"
        ws.column_dimensions["A"].width = 11
        for j in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 10.5

    # 안내 시트
    ws = wb.create_sheet("안내", 0)
    info = [
        ("파일 설명", "코스피/코스닥 현물, 선물(KOSPI200) 투자자별 순매수 동향 일일 트래킹"),
        ("단위", "순매수 금액, 억원 (지수 종가 제외)"),
        ("출처", "네이버 금융 투자자별 매매동향 (finance.naver.com)"),
        ("수집 시작일", BACKFILL_START.strftime("%Y-%m-%d")),
        ("누적 컬럼", "수집 시작일부터의 연초 누적 순매수 (YTD)"),
        ("마지막 갱신", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("주의", "이 파일은 collect.py 가 매일 자동 재생성 — 직접 편집 금지"),
        ("색상", "빨강 = 순매수(+), 파랑 = 순매도(-)"),
    ]
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 70
    for i, (k, v) in enumerate(info, start=2):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=v)
        a.font = Font(name="맑은 고딕", size=10, bold=True)
        b.font = f_base

    wb.save(XLSX_PATH)
    logging.info("엑셀 저장: %s", XLSX_PATH)


def build_dashboard_data(flow_dfs, index_dfs):
    payload = {
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "unit": "억원",
        "markets": {},
    }
    for mkey, _, mname, icode in MARKETS:
        df = flow_dfs[mkey].copy()
        idx = index_dfs[mkey]
        if not idx.empty:
            df = df.merge(idx, on="date", how="left")
        else:
            df["close"] = None
        m = {
            "name": mname,
            "indexName": {"KOSPI": "코스피", "KOSDAQ": "코스닥", "KPI200": "KOSPI200"}[icode],
            "dates": [d.strftime("%Y-%m-%d") for d in df["date"]],
            "close": [None if pd.isna(v) else round(float(v), 2) for v in df["close"]],
        }
        for c in FLOW_COLS:
            m[c] = [int(v) for v in df[c]]
        payload["markets"][mkey] = m
    js = "window.FLOW_DATA = " + json.dumps(payload, ensure_ascii=False) + ";\n"
    with open(DASH_JS_PATH, "w", encoding="utf-8") as f:
        f.write(js)
    logging.info("대시보드 데이터 저장: %s", DASH_JS_PATH)


def main():
    setup_logging()
    backfill = "--backfill" in sys.argv
    os.makedirs(DATA_DIR, exist_ok=True)

    flow_dfs, index_dfs = {}, {}
    for mkey, sosok, mname, icode in MARKETS:
        csv_path = os.path.join(DATA_DIR, f"{mkey}.csv")
        old = load_csv(csv_path)
        if backfill or old.empty:
            start = BACKFILL_START
        else:
            start = max(old["date"]) - timedelta(days=14)  # 최근 잠정치 갱신
        new_rows = fetch_flows(sosok, start, mname)
        df = upsert(old, new_rows, FLOW_COLS)
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        flow_dfs[mkey] = df

        icsv = os.path.join(DATA_DIR, f"index_{mkey}.csv")
        iold = load_csv(icsv)
        istart = BACKFILL_START if (backfill or iold.empty) else max(iold["date"]) - timedelta(days=14)
        try:
            idx_rows = fetch_index(icode, istart)
            idf = upsert(iold, idx_rows, ["close", "change_pct"])
            idf.to_csv(icsv, index=False, encoding="utf-8-sig")
        except Exception as e:  # noqa: BLE001
            logging.warning("%s 지수 수집 실패(수급만 저장): %s", icode, e)
            idf = iold
        index_dfs[mkey] = idf

    build_excel(flow_dfs, index_dfs)
    build_dashboard_data(flow_dfs, index_dfs)

    last = max(flow_dfs["kospi"]["date"])
    logging.info("완료 — 최신 거래일 %s, 코스피 %d일 / 코스닥 %d일 / 선물 %d일",
                 last, len(flow_dfs["kospi"]), len(flow_dfs["kosdaq"]), len(flow_dfs["futures"]))


if __name__ == "__main__":
    main()
