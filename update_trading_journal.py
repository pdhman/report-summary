# -*- coding: utf-8 -*-
"""
update_trading_journal.py
거래소 CSV 히스토리(Time, Coin, Direction, Price, Size, Trade Value, Fee, Fee Token, Closed PNL)를
매매일지 엑셀의 'trading' 시트에 자동 반영하는 스크립트.

사용법:
    python update_trading_journal.py <매매일지.xlsx> <히스토리.csv> [--since YYYY-MM-DD] [--out 출력.xlsx]

동작 방식:
  1) CSV를 시간순으로 정렬해 코인별 포지션을 재구성한다.
     - Open이 쌓이면 가중평균 진입가 갱신, Close로 수량이 0이 되면 '완결 매매' 1건으로 집계
     - 부분 청산은 청산가를 가중평균으로 합산, Closed PNL은 전부 합산
     - 'Liquidation'이 포함된 청산은 강제청산으로 표시
  2) 시트의 마지막 기록 날짜(--since 미지정 시 자동 감지) 이후의 완결 매매만 새 행으로 추가한다.
  3) 보유중 포지션은 시트에서 (종목, 롱/숏, 청산가 비어있음) 행을 찾아 진입가/수량을 갱신하고,
     없으면 '포지션 보유중' 메모와 함께 새 행을 추가한다.
     완결된 매매가 기존 '보유중' 행과 매칭되면 새 행 대신 그 행의 청산가/손익을 채운다.

주의: 이 스크립트는 값만 기록하며 기존 수식/서식은 건드리지 않는다.
      엑셀에서 파일을 열면 다른 시트의 수식은 자동 재계산된다.
"""
import argparse
import copy
import csv
import datetime as dt
import io
import re
import sys

import openpyxl

SHEET = 'trading'
SYMBOL_MAP = {'WTIOIL': 'WTI'}          # CSV 심볼 -> 일지 표기
COLS = 'ABCDEFGHIJKLMNOPQ'

# ---------------------------------------------------------------- CSV 파싱
TIME_RE = re.compile(r'(\d{4})\.\s*(\d+)\.\s*(\d+)\.\s*(\S*)\s+(\d+):(\d+):(\d+)')

def parse_time(s):
    """'2026. 8. 21. 오전 10:20:05' 형태(인코딩 깨짐 포함)와 ISO 형태 모두 지원."""
    s = s.strip()
    m = TIME_RE.match(s)
    if m:
        y, mo, d, ampm, h, mi, se = m.groups()
        h = int(h)
        if '후' not in ampm and '전' not in ampm:
            raise ValueError(f'오전/오후를 판별할 수 없습니다(인코딩 확인 필요): {s!r}')
        if '후' in ampm and h != 12:      # 오후
            h += 12
        if '전' in ampm and h == 12:      # 오전 12시 = 0시
            h = 0
        return dt.datetime(int(y), int(mo), int(d), h, int(mi), int(se))
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S'):
        try:
            return dt.datetime.strptime(s[:19], fmt)
        except ValueError:
            pass
    raise ValueError(f'시간 형식을 해석할 수 없습니다: {s!r}')


def read_csv(path):
    raw = open(path, 'rb').read()
    candidates = []
    for enc, err in (('utf-8-sig', 'strict'), ('utf-8', 'replace'), ('cp949', 'replace')):
        try:
            candidates.append(raw.decode(enc, errors=err))
        except UnicodeDecodeError:
            pass
    def score(t):
        # 오전/오후가 온전히 읽히는 디코딩을 선호 (바이트 일부가 깨진 파일 대비)
        return t.count('전 ') + t.count('후 ')
    txt = max(candidates, key=score)
    rows = list(csv.DictReader(io.StringIO(txt)))
    events = []
    for r in rows:
        if not r.get('Time'):
            continue
        sym = SYMBOL_MAP.get(r['Coin'].strip(), r['Coin'].strip())
        events.append({
            'time': parse_time(r['Time']),
            'sym': sym,
            'dir': r['Direction'].strip(),
            'price': float(r['Price']),
            'size': float(r['Size']),
            'pnl': float(r['Closed PNL'] or 0),
        })
    events.sort(key=lambda e: e['time'])
    return events

# ------------------------------------------------------- 포지션 재구성
def rebuild_trades(events):
    """완결 매매 리스트와 보유중 포지션 dict를 반환."""
    pos = {}      # sym -> state
    closed = []
    orphans = []  # 대응 Open이 없는 청산(기간 밖 진입)
    for e in events:
        d = e['dir']
        is_open = d.startswith('Open')
        is_close = 'Close' in d
        side = '롱' if 'Long' in d else '숏'
        st = pos.get(e['sym'])
        if is_open:
            if st is None:
                st = pos[e['sym']] = {
                    'side': side, 'size': 0.0, 'avg': 0.0,
                    'open_time': e['time'], 'closes': [], 'pnl': 0.0,
                    'max_size': 0.0, 'entries': [], 'liq': False,
                }
            new_size = st['size'] + e['size']
            st['avg'] = (st['avg'] * st['size'] + e['price'] * e['size']) / new_size
            st['size'] = new_size
            st['max_size'] = max(st['max_size'], new_size)
            st['entries'].append((e['time'], e['price'], e['size']))
        elif is_close and st is None:
            orphans.append(e)
        elif is_close and st is not None:
            st['closes'].append((e['time'], e['price'], e['size']))
            st['pnl'] += e['pnl']
            st['size'] = max(st['size'] - e['size'], 0.0)
            if 'Liquidation' in d:
                st['liq'] = True
            if st['size'] < 1e-9:                       # 포지션 종료 -> 매매 1건
                tot = sum(s for _, _, s in st['closes'])
                exit_avg = sum(p * s for _, p, s in st['closes']) / tot
                entry_tot = sum(s for _, _, s in st['entries'])
                entry_avg = sum(p * s for _, p, s in st['entries']) / entry_tot
                closed.append({
                    'sym': e['sym'], 'side': st['side'],
                    'open_time': st['open_time'], 'close_time': e['time'],
                    'entry': entry_avg, 'exit': exit_avg,
                    'size': st['max_size'], 'pnl': st['pnl'],
                    'notional': entry_avg * st['max_size'],
                    'n_entries': len(st['entries']), 'n_closes': len(st['closes']),
                    'liq': st['liq'],
                })
                del pos[e['sym']]
    return closed, pos, orphans

# ------------------------------------------------------- 엑셀 반영
def style_like(ws, src_row, dst_row):
    for col in COLS:
        s, t = ws[f'{col}{src_row}'], ws[f'{col}{dst_row}']
        t.font = copy.copy(s.font)
        t.border = copy.copy(s.border)
        t.fill = copy.copy(s.fill)
        t.alignment = copy.copy(s.alignment)
        t.number_format = s.number_format


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('csv')
    ap.add_argument('--since', help='이 날짜(YYYY-MM-DD) 이후에 청산된 매매만 추가. 미지정 시 시트의 마지막 날짜 자동 사용')
    ap.add_argument('--out', help='다른 파일명으로 저장(미지정 시 원본 덮어쓰기)')
    args = ap.parse_args()

    events = read_csv(args.csv)
    closed, open_pos, orphans = rebuild_trades(events)
    if orphans:
        print(f'주의: 대응 진입이 CSV에 없는 청산 {len(orphans)}건은 건너뜀 '
              f'(더 긴 기간의 히스토리로 다시 실행 권장)')

    wb = openpyxl.load_workbook(args.xlsx)
    if SHEET not in wb.sheetnames:
        sys.exit(f"'{SHEET}' 시트를 찾을 수 없습니다.")
    ws = wb[SHEET]

    # 기존 기록 스캔
    last_date = None
    open_rows = {}          # (sym, side) -> row  (청산가 비어있는 행)
    seen = []               # 이미 기록된 완결 매매 [(종목, 진입가, 청산가, 손익)]
    last_data_row = 1
    for r in range(2, ws.max_row + 1):
        b = ws[f'B{r}'].value
        if b is None:
            continue
        last_data_row = r
        a = ws[f'A{r}'].value
        if isinstance(a, dt.datetime) and (last_date is None or a > last_date):
            last_date = a
        if ws[f'H{r}'].value in (None, '') and ws[f'G{r}'].value not in (None, ''):
            open_rows[(str(b).strip(), str(ws[f'C{r}'].value or '').strip())] = r
        else:
            try:
                seen.append((str(b).strip(),
                             float(ws[f'G{r}'].value),
                             float(ws[f'H{r}'].value),
                             float(ws[f'L{r}'].value) if isinstance(ws[f'L{r}'].value, (int, float)) else None))
            except (TypeError, ValueError):
                pass

    since = dt.datetime.strptime(args.since, '%Y-%m-%d') if args.since else (last_date or dt.datetime.min)
    template_row = last_data_row if last_data_row > 1 else 2
    added, filled, updated = 0, 0, 0

    def next_row():
        nonlocal last_data_row
        last_data_row += 1
        style_like(ws, template_row, last_data_row)
        return last_data_row

    # 1) 완결 매매 반영
    for t in sorted(closed, key=lambda x: x['close_time']):
        key = (t['sym'], t['side'])
        note_parts = []
        if t['n_entries'] > 1:
            note_parts.append(f"{t['n_entries']}회 분할 진입(평균 {t['entry']:.2f})")
        if t['n_closes'] > 1:
            note_parts.append(f"{t['n_closes']}회 분할 청산(평균 {t['exit']:.2f})")
        if t['liq']:
            note_parts.append('강제청산(Liquidation)')
        note = ', '.join(note_parts)

        def close_enough(a, b, tol=0.005):
            return a is not None and b is not None and abs(a - b) <= tol * max(abs(b), 1)
        if any(s[0] == t['sym'] and
               (close_enough(t['entry'], s[1]) and close_enough(t['exit'], s[2])
                or (s[3] is not None and abs(t['pnl'] - s[3]) < 1.0))
               for s in seen):
            continue                               # 이미 일지에 있는 매매
        # '보유중' 행 완결 처리는 진입가가 실제로 일치할 때만 수행
        if key in open_rows and close_enough(t['entry'], ws[f'G{open_rows[key]}'].value, 0.01):
            r = open_rows.pop(key)
            ws[f'H{r}'] = round(t['exit'], 2)
            ws[f'L{r}'] = round(t['pnl'], 2)
            ws[f'M{r}'] = round(t['pnl'] / t['notional'], 4)
            if note:
                ws[f'P{r}'] = note
            filled += 1
        elif t['close_time'] > since:             # 신규 행 추가
            r = next_row()
            ws[f'A{r}'] = t['open_time'].replace(hour=0, minute=0, second=0)
            ws[f'B{r}'] = t['sym']
            ws[f'C{r}'] = t['side']
            ws[f'F{r}'] = round(t['size'], 4)
            ws[f'G{r}'] = round(t['entry'], 2)
            ws[f'H{r}'] = round(t['exit'], 2)
            ws[f'L{r}'] = round(t['pnl'], 2)
            ws[f'M{r}'] = round(t['pnl'] / t['notional'], 4)
            if note:
                ws[f'P{r}'] = note
            added += 1

    # 2) 보유중 포지션 반영
    for sym, st in open_pos.items():
        key = (sym, st['side'])
        if key in open_rows:                      # 기존 행 갱신(추가 진입 반영)
            r = open_rows[key]
            ws[f'F{r}'] = round(st['size'], 4)
            ws[f'G{r}'] = round(st['avg'], 2)
            updated += 1
        else:
            r = next_row()
            ws[f'A{r}'] = st['open_time'].replace(hour=0, minute=0, second=0)
            ws[f'B{r}'] = sym
            ws[f'C{r}'] = st['side']
            ws[f'F{r}'] = round(st['size'], 4)
            ws[f'G{r}'] = round(st['avg'], 2)
            ws[f'P{r}'] = f"포지션 보유중 — {len(st['entries'])}회 진입(평균 {st['avg']:.2f})"
            added += 1

    out = args.out or args.xlsx
    wb.save(out)
    print(f'신규 행 추가: {added}건 / 보유중 행 완결 처리: {filled}건 / 보유중 행 갱신: {updated}건')
    print(f'기준일(이후 청산분만 추가): {since:%Y-%m-%d} / 저장: {out}')


if __name__ == '__main__':
    main()
